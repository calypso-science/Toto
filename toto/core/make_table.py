from openpyxl import Workbook,load_workbook
import sys
import numpy as np
from openpyxl.styles import Font
from openpyxl.styles import Alignment

def create_table(dest_filename,sheetname,mat):

    try: # try loading existing excel file
        wb = load_workbook(dest_filename)
    except:
        wb = Workbook()


    # Create an new worksheet
    ws = wb.create_sheet(title=sheetname)
    
    # delete Sheet
    if 'Sheet' in wb.get_sheet_names():
        std=wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)
 
    ##get Font for the title
    font_title = Font(name='Open Sans',size=10,bold=True)
    Al=Alignment(horizontal='center',vertical='center')
    reste_title = Font(name='Open Sans',size=10)

    # Write title.
    for a in range(0,mat.shape[1]):
        ws.cell(column=a+1,row=1,value="%s" % mat[0,a]).font=font_title
        ws.cell(column=a+1,row=1).alignment=Al


    for b in range(0,mat.shape[0]):
        ws.cell(column=1,row=b+1,value="%s" % mat[b,0]).font=font_title
        ws.cell(column=1,row=b+1).alignment=Al

    for a in range(1,mat.shape[1]):
        for b in range(1,mat.shape[0]):
            ws.cell(column=a+1,row=b+1,value="%s" % mat[b,a]).font=reste_title
            ws.cell(column=a+1,row=b+1).alignment=Al



    # save the excel file

    wb.save(filename = dest_filename)


##############################################################

if __name__ == '__main__':
    filename = str(sys.argv[1])   
    sheetname = str(sys.argv[2])
    mat = np.array(eval(sys.argv[3]))
    create_table(filename,sheetname,mat)
