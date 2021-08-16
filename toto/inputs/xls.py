"""Read xls,xlsx file
    This import Excel type file. The function uses the read_excel function from panda <https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html>_.
    This class returns a Panda Dataframe with some extra attributes such as Latitude,Longitude,Units.
    
    Parameters
    ~~~~~~~~~~

    filename : (files,) str or list_like
        A list of filename to process.
    sheet_name : str, int, list, or None, default 0
        Strings are used for sheet names. Integers are used in zero-indexed
        sheet positions. Lists of strings/integers are used to request
        multiple sheets. Specify None to get all sheets.

        Available cases:

        * Defaults to ``0``: 1st sheet as a `DataFrame`
        * ``1``: 2nd sheet as a `DataFrame`
        * ``"Sheet1"``: Load sheet with name "Sheet1"
        * ``[0, 1, "Sheet5"]``: Load first, second and sheet named "Sheet5"
          as a dict of `DataFrame`
        * None: All sheets.
    colNames : List, default []
        List of column names to use.
    unitNames : List, default []
        List of unit to use. 
    miss_val : scalar, str, list-like, or dict, optional
        Additional strings to recognize as NA/NaN. If dict passed, specific
        per-column NA values. 
    colNamesLine : int, default 1
        Line number where the header are defined
    skiprows : list-like, int or callable, optional
        Line numbers to skip (0-indexed) or number of lines to skip (int)
        at the start of the file.
        If callable, the callable function will be evaluated against the row
        indices, returning True if the row should be skipped and False otherwise.
        An example of a valid callable argument would be ``lambda x: x in [0, 2]``.
    skipfooter : int, default 0
        Number of lines at bottom of file to skip (Unsupported with engine='c').
    unitNamesLine : int, default 1
        Line number where the units are defined 
    single_column : bool, default False
        The time is represented in a single column
    customUnit : str, default '%d-%m-%Y %H:%M:%S'
        String reprensenting the time format
    unit : str default 's', can be 'auto','custom','matlab' or 's' and 'D'
        unit of the single column time. Only matter if `single_column` is True
    time_col_name: dict, default {'Year':'year','Month':'month','Day':'day','Hour':'hour','Min':'Minute','Sec':'Second'}
        Dictonary for renaming the each column, so Panda can interprate the time. Only matter if `single_column` is False
    

    Examples
    ~~~~~~~~
    
    >>> from toto.inputs.xls import XLSfile
    >>> tx=XLSfile([filename],sheetnames='test3', colNames= [], unitNames= [],miss_val='NaN', colNamesLine= 1, skiprows= 2, unitNamesLine= 0,\
    skipfooter= 0, single_column= True, unit= 's',\
    customUnit= '%d-%m-%Y %H:%M:%S', time_col_name= {})
    >>> tx.reads()
    >>> tx.read_time()
    >>> df=tx._toDataFrame()

"""
import glob,os,sys
import pandas as pd
import datetime as dt
import numpy as np
from openpyxl import load_workbook
_NUMERIC_KINDS = set('buifc')


class XLSfile():

    @staticmethod
    def defaultExtensions():
        return ['.xls','.xlsx']


    def __init__(self,filename,sheetnames=[],\
                               colNames=[],\
                               unitNames=[],\
                               miss_val='NaN',\
                               colNamesLine=1,\
                               skiprows=0,\
                               unitNamesLine=0,\
                               skipfooter=0,\
                               single_column=False,\
                               unit='s',\
                               customUnit='%d-%m-%Y %H:%M:%S',\
                               time_col_name={'Year':'year','Month':'month','Day':'day','Hour':'hour','Min':'Minute','Sec':'Second'},\
                               ):
        self.filename     = filename
        self.fileparam=lambda:None
        self.fileparam.sheetnames=sheetnames
        self.fileparam.colNames     = colNames
        self.fileparam.unitNames     = unitNames
        self.fileparam.miss_val     = miss_val
        self.fileparam.ext=self.defaultExtensions()

        self.fileparam.colNamesLine = colNamesLine
        self.fileparam.skiprows     = skiprows
        self.fileparam.unitNamesLine = unitNamesLine
        self.fileparam.skipfooter = skipfooter

        self.fileparam.single_column=single_column
        self.fileparam.unit=unit
        self.fileparam.customUnit=customUnit


        self.fileparam.time_col_name=time_col_name

        self.encoding    = None

        self.data=[]

        # set usr defined parameter

    def reads(self):
        for file in self.filename:
            self.read(file)

    def read(self,filename):
        wb = load_workbook(filename)
        if self.fileparam.sheetnames==[]:
            self.fileparam.sheetnames=list(wb.keys())

        if isinstance(self.fileparam.sheetnames,str):
            self.fileparam.sheetnames=[self.fileparam.sheetnames]

        
        # # unit header
        if self.fileparam.unitNamesLine:
            for i,sheet in enumerate(self.fileparam.sheetnames):
                df=wb[sheet]
                unit={}
                unitNames=[x.value for x in df[self.fileparam.unitNamesLine] if x.value]
                for i,col in enumerate(self.fileparam.colNames): 
                    unit[col]=unitNames[i].replace('[','').replace(']','')
                self.fileparam.unitNames.append(unit)


        for i,sheet in enumerate(self.fileparam.sheetnames):
            df=wb[sheet]
            colname=[x.value for x in df[self.fileparam.colNamesLine] if x.value]
            self.fileparam.colNames.append(colname)
            try:
                df=pd.read_excel(filename,sheet_name=sheet,skiprows=self.fileparam.skiprows,\
                        header=None,names=colname,skipfooter=self.fileparam.skipfooter,\
                        na_values=self.fileparam.miss_val,engine='openpyxl')
            except pd.errors.ParserError as e:
                raise WrongFormatError('XLS File {}: '.format(filename)+e.args[0])

            keys=list(df.keys())
            for key in keys:
                if np.all(df[key].isna()):
                    del df[key] 

                # ### check if all numerics
                # if not np.asarray(df[key].values).dtype.kind in _NUMERIC_KINDS:
                #     print('Warning %s: this key was deleted containes string' % key)
                #     del df[key]

            self.data.append(df)

    def read_time(self):

        for i,df in enumerate(self.data):

            if self.fileparam.single_column is True:
                if self.fileparam.unit == 'auto':
                    time=pd.to_datetime(df[self.fileparam.time_col_name])
                elif self.fileparam.unit == 's' or self.fileparam.unit == 'D':
                    time=pd.to_datetime(df[self.fileparam.time_col_name],unit=self.fileparam.unit)
                elif self.fileparam.unit == 'custom':
                    time=pd.to_datetime(df[self.fileparam.time_col_name],format=self.fileparam.customUnit)
                elif self.fileparam.unit == 'excel':
                    time=pd.TimedeltaIndex(df[self.fileparam.time_col_name], unit='d') + dt.datetime(1970,1,1)
                    time=pd.to_datetime(time)
                    del df[self.fileparam.time_col_name]
            else:
                old_name=self.fileparam.time_col_name.keys()
                time=pd.to_datetime(df[old_name].rename(columns=self.fileparam.time_col_name))

                for oldkey in old_name:
                    del df[oldkey]



            self.data[i]=df
            self.data[i]['time']=time
            print(time)
            self.data[i].set_index('time',inplace=True,drop=False)
            

            self.add_unit(i)
            keys=self.data[i].keys()
            for key in keys:
                self.data[i][key].long_name=key  
    def add_unit(self,i):

        keys=self.data[i].keys()
        for key in keys:
            if key in self.fileparam.unitNames:
                units=self.fileparam.unitNames[key]
                self.data[i][key].units=units

            elif '[' in key and ']' in key:
                a,b=key.split('[')
                self.data[i].rename(columns={key: a},inplace=True)
                self.data[i][a].units=b.split(']')[0]



    def _toDataFrame(self):
        return self.data
